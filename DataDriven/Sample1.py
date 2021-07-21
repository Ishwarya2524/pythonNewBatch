import openpyxl
excel_loc=r"C:\Users\ARUL KUMARAN\PycharmProjects\pythonProjectselenium10.00 PM\Excel\New Microsoft Office Excel Worksheet.xlsx"
w=openpyxl.load_workbook(excel_loc)
sheet=w.active
for row_no in range(1,sheet.max_row +1):
    for cell_no in range(1,sheet.max_column +1):
        c=sheet.cell(row_no,cell_no)
        data=c.value
        print(data)
